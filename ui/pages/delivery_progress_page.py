# app/ui/pages/delivery_progress_page.py
import streamlit as st
import pandas as pd
from datetime import date, timedelta, datetime
from typing import Dict, Optional, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class DeliveryProgressPage:
    """ç´å…¥é€²åº¦ç®¡ç†ãƒšãƒ¼ã‚¸"""
    
    def __init__(self, transport_service, auth_service=None):
        self.service = transport_service
        self.auth_service = auth_service

    def _can_edit_page(self) -> bool:
        """ãƒšãƒ¼ã‚¸ç·¨é›†æ¨©é™ãƒã‚§ãƒƒã‚¯"""
        if not self.auth_service:
            return True
        user = st.session_state.get('user')
        if not user:
            return False
        return self.auth_service.can_edit_page(user['id'], "ç´å…¥é€²åº¦")
    
    def show(self):
        """ãƒšãƒ¼ã‚¸è¡¨ç¤º"""
        st.title("ğŸ“‹ ç´å…¥é€²åº¦ç®¡ç†")
        st.write("å—æ³¨ã‹ã‚‰å‡ºè·ã¾ã§ã®é€²æ—ã‚’ç®¡ç†ã—ã¾ã™ã€‚")

        # æ¨©é™ãƒã‚§ãƒƒã‚¯
        can_edit = self._can_edit_page()
        if not can_edit:
            st.warning("âš ï¸ ã“ã®ç”»é¢ã®ç·¨é›†æ¨©é™ãŒã‚ã‚Šã¾ã›ã‚“ã€‚é–²è¦§ã®ã¿å¯èƒ½ã§ã™ã€‚")

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "ğŸ“Š é€²åº¦ä¸€è¦§",
            "âœ… å®Ÿç¸¾ç™»éŒ²",
            "â• æ–°è¦ç™»éŒ²",
            "ğŸ“¦ å‡ºè·å®Ÿç¸¾",
            "ğŸ­ ç¤¾å†…æ³¨æ–‡"
        ])

        with tab1:
            self._show_progress_list(can_edit)
        with tab2:
            self._show_actual_registration(can_edit)
        with tab3:
            self._show_progress_registration(can_edit)
        with tab4:
            self._show_shipment_records()
        with tab5:
            self._show_internal_orders()
    
    def _show_progress_list(self, can_edit):
        """é€²åº¦ä¸€è¦§è¡¨ç¤º"""
        st.header("ğŸ“Š ç´å…¥é€²åº¦ä¸€è¦§")
        
        # ã‚µãƒãƒªãƒ¼è¡¨ç¤º
        try:
            summary = self.service.get_progress_summary()
            
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("ç·ã‚ªãƒ¼ãƒ€ãƒ¼æ•°", summary.get('total_orders', 0))
            with col2:
                st.metric("æœªå‡ºè·", summary.get('unshipped', 0))
            with col3:
                st.metric("ä¸€éƒ¨å‡ºè·", summary.get('partial', 0))
            with col4:
                st.metric("é…å»¶", summary.get('delayed', 0), delta_color="inverse")
            with col5:
                st.metric("ç·Šæ€¥", summary.get('urgent', 0), delta_color="inverse")
        
        except Exception as e:
            st.warning(f"ã‚µãƒãƒªãƒ¼å–å¾—ã‚¨ãƒ©ãƒ¼: {e}")

        # ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼ - ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆã‚’éå»10æ—¥é–“ã«å¤‰æ›´
        st.subheader("ğŸ” ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼")
        col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)

        with col_f1:
            start_date = st.date_input(
                "ç´æœŸï¼ˆé–‹å§‹ï¼‰",
                value=date.today() - timedelta(days=1),
                key="progress_start_date"
            )

        with col_f2:
            end_date = st.date_input(
                "ç´æœŸï¼ˆçµ‚äº†ï¼‰",
                value=date.today()+timedelta(days=10),
                key="progress_end_date"
            )

        with col_f3:
            status_filter = st.multiselect(
                "ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹",
                options=['æœªå‡ºè·', 'è¨ˆç”»æ¸ˆ', 'ä¸€éƒ¨å‡ºè·', 'å‡ºè·å®Œäº†'],
                default=['æœªå‡ºè·', 'è¨ˆç”»æ¸ˆ', 'ä¸€éƒ¨å‡ºè·', 'å‡ºè·å®Œäº†'],
                key="progress_status_filter"
            )

        with col_f4:
            product_filter = st.text_input(
                "è£½å“ã‚³ãƒ¼ãƒ‰ï¼ˆéƒ¨åˆ†ä¸€è‡´ï¼‰",
                key="progress_product_filter"
            ).strip()

        with col_f5:
            # è£½å“ç¾¤ãƒ•ã‚£ãƒ«ã‚¿ã‚’è¿½åŠ 
            try:
                product_groups_df = self.service.product_repo.get_all_product_groups(include_inactive=False)
                if not product_groups_df.empty:
                    group_options = ['å…¨ã¦'] + product_groups_df['group_name'].dropna().unique().tolist()
                    product_group_filter = st.selectbox(
                        "è£½å“ç¾¤",
                        options=group_options,
                        key="progress_product_group_filter"
                    )
                else:
                    product_group_filter = 'å…¨ã¦'
            except Exception as e:
                st.warning(f"è£½å“ç¾¤ãƒ‡ãƒ¼ã‚¿å–å¾—ã‚¨ãƒ©ãƒ¼: {e}")
                product_group_filter = 'å…¨ã¦'
        # é€²åº¦ãƒ‡ãƒ¼ã‚¿å–å¾—
        try:
            progress_df = self.service.get_delivery_progress(start_date, end_date)

            # â–¼ å®Ÿç¸¾é€²åº¦ï¼ˆshipped_remaining_quantityï¼‰ã®å†è¨ˆç®—
            with st.expander("å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—ï¼ˆshipped_remaining_quantityï¼‰"):
                # è£½å“ãƒªã‚¹ãƒˆã‚’å–å¾—
                try:
                    products = self.service.product_repo.get_all_products()
                    if not products.empty:
                        sr_product_options = {
                            f"{row['product_code']} - {row['product_name']}": row['id']
                            for _, row in products.iterrows()
                        }
                        sr_selected_product = st.selectbox(
                            "è£½å“ã‚³ãƒ¼ãƒ‰",
                            options=list(sr_product_options.keys()),
                            key="sr_product_select"
                        )
                        sr_product_id = sr_product_options[sr_selected_product]
                    else:
                        st.warning("è£½å“ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
                        sr_product_id = None
                except:
                    st.error("è£½å“ãƒ‡ãƒ¼ã‚¿å–å¾—ã‚¨ãƒ©ãƒ¼")
                    sr_product_id = None

                sr_start_date = st.date_input("å†è¨ˆç®—é–‹å§‹æ—¥ï¼ˆå®Ÿç¸¾ï¼‰", key="sr_start_date")
                sr_end_date = st.date_input("å†è¨ˆç®—çµ‚äº†æ—¥ï¼ˆå®Ÿç¸¾ï¼‰", key="sr_end_date")

                col_sr_one, col_sr_all = st.columns(2)

                with col_sr_one:
                    if st.button("é¸æŠè£½å“ã®å®Ÿç¸¾é€²åº¦ã‚’å†è¨ˆç®—", key="btn_sr_one", disabled=not can_edit):
                        if sr_product_id:
                            self.service.recompute_shipped_remaining(sr_product_id, sr_start_date, sr_end_date)
                            st.success("å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—ãŒå®Œäº†ã—ã¾ã—ãŸ")
                        else:
                            st.error("è£½å“ã‚’é¸æŠã—ã¦ãã ã•ã„")

                with col_sr_all:
                    if st.button("å…¨è£½å“ã®å®Ÿç¸¾é€²åº¦ã‚’å†è¨ˆç®—", key="btn_sr_all", disabled=not can_edit):
                        self.service.recompute_shipped_remaining_all(sr_start_date, sr_end_date)
                        st.success("å…¨è£½å“ã®å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—ãŒå®Œäº†ã—ã¾ã—ãŸ")
                              
            if not progress_df.empty:
                # ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼é©ç”¨
                if status_filter:
                    progress_df = progress_df[progress_df['status'].isin(status_filter)]
                if product_filter:
                    progress_df = progress_df[
                        progress_df['product_code'].fillna('').str.contains(product_filter, case=False, na=False)
                    ]
                # è£½å“ç¾¤ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼é©ç”¨
                if product_group_filter and product_group_filter != 'å…¨ã¦':
                    progress_df = progress_df[
                        progress_df['product_group_name'].fillna('') == product_group_filter
                    ]
                
                # è¡¨ç¤ºå½¢å¼é¸æŠã‚’è¿½åŠ 
                st.subheader("ğŸ“‹ è¡¨ç¤ºå½¢å¼")
                view_mode = st.radio(
                    "è¡¨ç¤ºãƒ¢ãƒ¼ãƒ‰",
                    options=['ä¸€è¦§è¡¨ç¤º', 'ãƒãƒˆãƒªãƒƒã‚¯ã‚¹è¡¨ç¤ºï¼ˆæ—¥ä»˜Ã—è£½å“ï¼‰'],
                    horizontal=True,
                    key="view_mode_selector"
                )
                
                if view_mode == 'ãƒãƒˆãƒªãƒƒã‚¯ã‚¹è¡¨ç¤ºï¼ˆæ—¥ä»˜Ã—è£½å“ï¼‰':
                    self._show_matrix_view(progress_df, can_edit)
                else:
                    # æ—¢å­˜ã®ä¸€è¦§è¡¨ç¤º
                    # ç·Šæ€¥åº¦ãƒ•ãƒ©ã‚°è¿½åŠ 
                    progress_df['days_to_delivery'] = (
                        pd.to_datetime(progress_df['delivery_date']) - pd.Timestamp(date.today())
                    ).dt.days
                    
                    progress_df['urgency'] = progress_df.apply(
                        lambda row: 'ğŸ”´é…å»¶' if row['days_to_delivery'] < 0 and row['status'] != 'å‡ºè·å®Œäº†'
                        else 'ğŸŸ¡ç·Šæ€¥' if 0 <= row['days_to_delivery'] <= 3 and row['status'] != 'å‡ºè·å®Œäº†'
                        else 'ğŸŸ¢',
                        axis=1
                    )
                    
                    # è¨ˆç”»é€²åº¦ã¨é€²åº¦ã‚’è¨ˆç®—
                    progress_df['planned_progress'] = (
                        progress_df.get('planned_quantity', 0).fillna(0) -
                        progress_df.get('order_quantity', 0).fillna(0)
                    )
                    progress_df['actual_progress'] = (
                        progress_df.get('shipped_quantity', 0).fillna(0) -
                        progress_df.get('order_quantity', 0).fillna(0)
                    )

                    st.subheader("ğŸ–Šï¸ æ‰‹å‹•è¨ˆç”»æ•°é‡ã®ä¸€æ‹¬ç·¨é›†")
                    # ä¸Šã®ä¸€è¦§è¡¨ç¤ºã¨åŒã˜åˆ—æ§‹æˆã«ã™ã‚‹ï¼ˆIDãªã—ï¼‰
                    editor_columns = ['urgency', 'product_code', 'product_name', 'delivery_date', 'order_quantity']

                    if 'manual_planning_quantity' in progress_df.columns:
                        editor_columns.append('manual_planning_quantity')
                    if 'planned_quantity' in progress_df.columns:
                        editor_columns.append('planned_quantity')

                    editor_columns.extend(['planned_progress', 'shipped_quantity', 'actual_progress', 'remaining_quantity', 'status'])

                    # IDã¯ä¿å­˜å‡¦ç†ã®ãŸã‚ã«åˆ¥é€”ä¿æŒ
                    editor_source = progress_df[editor_columns].copy()
                    editor_source.insert(0, 'id', progress_df['id'])
                    editor_source = editor_source.reset_index(drop=True)

                    original_editor = editor_source.copy()

                    # æ‰‹å‹•è¨ˆç”»ã®ã¿Float64å‹ã«å¤‰æ›ï¼ˆç·¨é›†å¯èƒ½ã«ã™ã‚‹ãŸã‚ï¼‰
                    if 'manual_planning_quantity' in editor_source.columns:
                        editor_source['manual_planning_quantity'] = editor_source['manual_planning_quantity'].astype('Float64')

                    # IDåˆ—ã‚’éè¡¨ç¤ºã«ã—ã¦12åˆ—ã§è¡¨ç¤º
                    edited_table = st.data_editor(
                        editor_source,
                        num_rows="fixed",
                        hide_index=True,
                        use_container_width=True,
                        column_config={
                            'id': None,  # IDåˆ—ã‚’éè¡¨ç¤º
                            'urgency': st.column_config.TextColumn('ç·Šæ€¥åº¦'),
                            'product_code': st.column_config.TextColumn('è£½å“ã‚³ãƒ¼ãƒ‰'),
                            'product_name': st.column_config.TextColumn('è£½å“å'),
                            'delivery_date': st.column_config.DateColumn('ç´æœŸ', format='YYYY-MM-DD'),
                            'order_quantity': st.column_config.NumberColumn('å—æ³¨æ•°', format='%d'),
                            'manual_planning_quantity': st.column_config.NumberColumn('æ‰‹å‹•è¨ˆç”»', min_value=0, step=1),
                            'planned_quantity': st.column_config.NumberColumn('è¨ˆç”»æ•°', format='%d'),
                            'planned_progress': st.column_config.NumberColumn('è¨ˆç”»é€²åº¦', format='%d'),
                            'shipped_quantity': st.column_config.NumberColumn('å‡ºè·æ¸ˆ', format='%d'),
                            'actual_progress': st.column_config.NumberColumn('é€²åº¦', format='%d'),
                            'remaining_quantity': st.column_config.NumberColumn('æ®‹æ•°', format='%d'),
                            'status': st.column_config.TextColumn('ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹'),
                        },
                        disabled=['urgency', 'product_code', 'product_name', 'delivery_date', 'order_quantity', 'planned_quantity', 'planned_progress', 'shipped_quantity', 'actual_progress', 'remaining_quantity', 'status'],
                        key="manual_plan_editor",
                    )
                    st.markdown(":red[æ‰‹å‹•è¨ˆç”»åˆ—ã®ã¿ç·¨é›†ã§ãã¾ã™ã€‚æ³¨æ„ï¼šï¼‘å¢—æ¸›åˆ†ã§ã¯ãªãã€å¤‰æ›´å¾Œã®æ•°å€¤ã‚’å…¥åŠ›ã€‚ï¼’ã€€ã‚­ãƒ£ãƒ³ã‚»ãƒ«ã™ã‚‹ã¨ãã€ï¼å…¥åŠ›ã§ã¯ãªãNoneãªã‚‹ã‚ˆã†ã«æ¶ˆã™]")

                    if st.button("æ‰‹å‹•è¨ˆç”»ã‚’ä¿å­˜", type="primary", key="save_manual_plans", disabled=not can_edit):
                        updated_count = 0
                        for idx, row in edited_table.iterrows():
                            new_val = row['manual_planning_quantity']
                            orig_val = original_editor.loc[idx, 'manual_planning_quantity']
                            if pd.isna(new_val) or new_val == '':
                                new_db_val = None
                            else:
                                try:
                                    new_db_val = int(new_val)
                                except (TypeError, ValueError):
                                    st.warning(f"ID {int(row['id'])} ã®å€¤ãŒç„¡åŠ¹ã§ã™ã€‚")
                                    continue

                            if pd.isna(orig_val):
                                orig_compare = None
                            else:
                                try:
                                    orig_compare = int(orig_val)
                                except (TypeError, ValueError):
                                    orig_compare = None

                            if orig_compare == new_db_val:
                                continue

                            success = self.service.update_delivery_progress(
                                int(row['id']),
                                {'manual_planning_quantity': new_db_val}
                            )
                            if success:
                                updated_count += 1
                            else:
                                st.error(f"ID {int(row['id'])} ã®æ›´æ–°ã«å¤±æ•—ã—ã¾ã—ãŸã€‚")

                        if updated_count:
                            st.success(f"{updated_count} ä»¶ã®æ‰‹å‹•è¨ˆç”»ã‚’æ›´æ–°ã—ã¾ã—ãŸã€‚")
                            st.rerun()
                        else:
                            st.info("å¤‰æ›´ã¯ã‚ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚")
                    
                    # è©³ç´°ç·¨é›†ãƒ»å‡ºè·å®Ÿç¸¾å…¥åŠ›
                    st.subheader("ğŸ“ è©³ç´°ç·¨é›†ãƒ»å‡ºè·å®Ÿç¸¾å…¥åŠ›")
                    
                    if not progress_df.empty:
                        # ã‚ªãƒ¼ãƒ€ãƒ¼é¸æŠ - è£½å“ã‚³ãƒ¼ãƒ‰è¡¨ç¤º
                        order_options = {
                            f"{row['order_id']} - {row['product_code']} ({row['delivery_date']})": row['id']
                            for _, row in progress_df.iterrows()
                        }
                        
                        selected_order_key = st.selectbox(
                            "ç·¨é›†ã™ã‚‹ã‚ªãƒ¼ãƒ€ãƒ¼ã‚’é¸æŠ",
                            options=list(order_options.keys()),
                            key="progress_edit_selector"
                        )
                        
                        if selected_order_key:
                            progress_id = order_options[selected_order_key]
                            progress_row = progress_df[progress_df['id'] == progress_id].iloc[0]
                            
                            # ã‚¿ãƒ–ã§ç·¨é›†ã¨å‡ºè·å®Ÿç¸¾ã‚’åˆ†é›¢
                            edit_tab, shipment_tab = st.tabs(["ğŸ“ é€²åº¦ç·¨é›†", "ğŸ“¦ å‡ºè·å®Ÿç¸¾å…¥åŠ›"])
                            
                            with edit_tab:
                                with st.form(f"edit_progress_{progress_id}"):
                                    st.write("**é€²åº¦æƒ…å ±ã‚’ç·¨é›†**")
                                    
                                    col_e1, col_e2 = st.columns(2)
                                    
                                    with col_e1:
                                        new_delivery_date = st.date_input(
                                            "ç´æœŸ",
                                            value=progress_row['delivery_date'],
                                            key=f"delivery_{progress_id}"
                                        )
                                        new_priority = st.number_input(
                                            "å„ªå…ˆåº¦ï¼ˆ1-10ï¼‰",
                                            min_value=1,
                                            max_value=10,
                                            value=int(progress_row.get('priority', 5)),
                                            key=f"priority_{progress_id}"
                                        )
                                    
                                    with col_e2:
                                        new_status = st.selectbox(
                                            "ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹",
                                            options=['æœªå‡ºè·', 'è¨ˆç”»æ¸ˆ', 'ä¸€éƒ¨å‡ºè·', 'å‡ºè·å®Œäº†', 'ã‚­ãƒ£ãƒ³ã‚»ãƒ«'],
                                            index=['æœªå‡ºè·', 'è¨ˆç”»æ¸ˆ', 'ä¸€éƒ¨å‡ºè·', 'å‡ºè·å®Œäº†', 'ã‚­ãƒ£ãƒ³ã‚»ãƒ«'].index(progress_row['status']) if progress_row['status'] in ['æœªå‡ºè·', 'è¨ˆç”»æ¸ˆ', 'ä¸€éƒ¨å‡ºè·', 'å‡ºè·å®Œäº†', 'ã‚­ãƒ£ãƒ³ã‚»ãƒ«'] else 0,
                                            key=f"status_{progress_id}"
                                        )
                                        new_notes = st.text_area(
                                            "å‚™è€ƒ",
                                            value=progress_row.get('notes', '') or '',
                                            key=f"notes_{progress_id}"
                                        )
                                    
                                    manual_value = progress_row.get('manual_planning_quantity')
                                    use_manual = st.checkbox(
                                        "æ‰‹å‹•è¨ˆç”»æ•°é‡ã‚’æŒ‡å®š",
                                        value=pd.notna(manual_value),
                                        key=f"use_manual_{progress_id}"
                                    )
                                    if pd.notna(manual_value):
                                        manual_default = int(manual_value)
                                    else:
                                        manual_default = int(progress_row.get('order_quantity', 0) or 0)
                                    manual_quantity = st.number_input(
                                        "æ‰‹å‹•è¨ˆç”»æ•°é‡",
                                        min_value=0,
                                        value=manual_default,
                                        step=1,
                                        key=f"manual_qty_{progress_id}",
                                        disabled=not use_manual
                                    )
                                    
                                    submitted = st.form_submit_button("ğŸ’¾ æ›´æ–°", type="primary", disabled=not can_edit)
                                    
                                    if submitted:
                                        update_data = {
                                            'delivery_date': new_delivery_date,
                                            'priority': new_priority,
                                            'status': new_status,
                                            'notes': new_notes,
                                            'manual_planning_quantity': int(manual_quantity) if use_manual else None
                                        }
                                        
                                        success = self.service.update_delivery_progress(progress_id, update_data)
                                        if success:
                                            st.success("é€²åº¦ã‚’æ›´æ–°ã—ã¾ã—ãŸ")
                                            st.rerun()
                                        else:
                                            st.error("é€²åº¦æ›´æ–°ã«å¤±æ•—ã—ã¾ã—ãŸ")
                            
                            # å‡ºè·å®Ÿç¸¾å…¥åŠ›ã‚¿ãƒ–
                            with shipment_tab:
                                # ç¾åœ¨ã®å‡ºè·çŠ¶æ³ã‚’è¡¨ç¤º
                                manual_display = progress_row.get('manual_planning_quantity')
                                manual_display = int(manual_display) if pd.notna(manual_display) else 'æœªè¨­å®š'
                                st.info(f"""
                                **ç¾åœ¨ã®çŠ¶æ³:**
                                - å—æ³¨æ•°: {progress_row.get('order_quantity', 0)}
                                - è¨ˆç”»æ•°: {progress_row.get('planned_quantity', 0)}
                                - æ‰‹å‹•è¨ˆç”»: {manual_display}
                                - å‡ºè·æ¸ˆ: {progress_row.get('shipped_quantity', 0)}
                                - æ®‹æ•°: {progress_row.get('remaining_quantity', 0)}
                                """)
                                
                                with st.form(f"shipment_form_{progress_id}"):
                                    st.write("**å‡ºè·å®Ÿç¸¾ã‚’å…¥åŠ›**")
                                    
                                    col_s1, col_s2 = st.columns(2)
                                    
                                    with col_s1:
                                        shipment_date = st.date_input(
                                            "å‡ºè·æ—¥ *",
                                            value=date.today(),
                                            key=f"ship_date_{progress_id}"
                                        )
                                        
                                        # ãƒˆãƒ©ãƒƒã‚¯é¸æŠ
                                        try:
                                            trucks_df = self.service.get_trucks()
                                            
                                            if not trucks_df.empty:
                                                truck_options = dict(zip(trucks_df['name'], trucks_df['id']))
                                                selected_truck = st.selectbox(
                                                    "ä½¿ç”¨ãƒˆãƒ©ãƒƒã‚¯ *",
                                                    options=list(truck_options.keys()),
                                                    key=f"ship_truck_{progress_id}"
                                                )
                                                truck_id = truck_options[selected_truck]
                                            else:
                                                st.warning("ãƒˆãƒ©ãƒƒã‚¯ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
                                                truck_id = None
                                        except:
                                            st.warning("ãƒˆãƒ©ãƒƒã‚¯æƒ…å ±ã®å–å¾—ã«å¤±æ•—ã—ã¾ã—ãŸ")
                                            truck_id = None
                                        
                                        remaining_qty = int(progress_row.get('remaining_quantity', 0))
                                        if remaining_qty > 0:
                                            shipped_quantity = st.number_input(
                                                "å‡ºè·æ•°é‡ *",
                                                min_value=1,
                                                max_value=remaining_qty,
                                                value=min(100, remaining_qty),
                                                key=f"ship_qty_{progress_id}"
                                            )
                                        else:
                                            st.warning("å‡ºè·å¯èƒ½ãªæ•°é‡ãŒã‚ã‚Šã¾ã›ã‚“")
                                            shipped_quantity = 0
                                    # delivery_progress_page.py ã®è©²å½“ç®‡æ‰€ã‚’ä¿®æ­£
                                    with col_s2:
                                        driver_name = st.text_input(
                                            "ãƒ‰ãƒ©ã‚¤ãƒãƒ¼å",
                                            key=f"driver_{progress_id}"
                                        )
                                        
                                        # ãƒˆãƒ©ãƒƒã‚¯ã®ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆæ™‚åˆ»ã‚’å–å¾—
                                        default_dep_time = None
                                        default_arr_time = None
                                        
                                        if truck_id and not trucks_df.empty:
                                            try:
                                                truck_row = trucks_df[trucks_df['id'] == truck_id]
                                                if not truck_row.empty:
                                                    truck_info = truck_row.iloc[0]
                                                    # departure_time ã¨ arrival_time ã‚«ãƒ©ãƒ ã‚’ä½¿ç”¨
                                                    if 'departure_time' in truck_info and pd.notna(truck_info['departure_time']):
                                                        default_dep_time = truck_info['departure_time']
                                                    if 'arrival_time' in truck_info and pd.notna(truck_info['arrival_time']):
                                                        default_arr_time = truck_info['arrival_time']
                                            except Exception as e:
                                                print(f"ãƒˆãƒ©ãƒƒã‚¯æ™‚åˆ»å–å¾—ã‚¨ãƒ©ãƒ¼: {e}")
                                        
                                        # ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆå€¤ã‚’è¨­å®š(ãƒˆãƒ©ãƒƒã‚¯è¨­å®šæ™‚åˆ»ãŒãªã‘ã‚Œã°None)
                                        actual_departure = st.time_input(
                                            "å®Ÿå‡ºç™ºæ™‚åˆ»",
                                            value=default_dep_time,
                                            key=f"dep_time_{progress_id}"
                                        )
                                        
                                        actual_arrival = st.time_input(
                                            "å®Ÿåˆ°ç€æ™‚åˆ»",
                                            value=default_arr_time,
                                            key=f"arr_time_{progress_id}"
                                        )
                                        
                                        shipment_notes = st.text_area(
                                            "å‚™è€ƒ",
                                            key=f"ship_notes_{progress_id}"
                                        )

                                    # å‡ºè·å®Ÿç¸¾ç™»éŒ²ãƒœã‚¿ãƒ³
                                    
                                    ship_submitted = st.form_submit_button("ğŸ“¦ å‡ºè·å®Ÿç¸¾ã‚’ç™»éŒ²", type="primary", disabled=not can_edit)
                                    
                                    if ship_submitted:
                                        if not truck_id:
                                            st.error("ãƒˆãƒ©ãƒƒã‚¯ã‚’é¸æŠã—ã¦ãã ã•ã„")
                                        elif shipped_quantity <= 0:
                                            st.error("å‡ºè·æ•°é‡ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„")
                                        else:
                                            shipment_data = {
                                                'progress_id': progress_id,
                                                'truck_id': truck_id,
                                                'shipment_date': shipment_date,
                                                'shipped_quantity': shipped_quantity,
                                                'driver_name': driver_name,
                                                'actual_departure_time': actual_departure,
                                                'actual_arrival_time': actual_arrival,
                                                'notes': shipment_notes
                                            }
                                            
                                            success = self.service.create_shipment_record(shipment_data)
                                            if success:
                                                st.success(f"âœ… å‡ºè·å®Ÿç¸¾ã‚’ç™»éŒ²ã—ã¾ã—ãŸï¼ˆ{shipped_quantity}å€‹ï¼‰")
                                                st.balloons()
                                                st.rerun()
                                            else:
                                                st.error("âŒ å‡ºè·å®Ÿç¸¾ç™»éŒ²ã«å¤±æ•—ã—ã¾ã—ãŸ")
                            
                            # å‰Šé™¤ãƒœã‚¿ãƒ³ã¯å¤–ã«é…ç½®
                            st.markdown("---")
                            col_del1, col_del2 = st.columns([1, 5])
                            with col_del1:
                                if st.button(f"ğŸ—‘ï¸ å‰Šé™¤", key=f"delete_progress_{progress_id}", type="secondary", disabled=not can_edit):
                                    success = self.service.delete_delivery_progress(progress_id)
                                    if success:
                                        st.success("é€²åº¦ã‚’å‰Šé™¤ã—ã¾ã—ãŸ")
                                        st.rerun()
                                    else:
                                        st.error("é€²åº¦å‰Šé™¤ã«å¤±æ•—ã—ã¾ã—ãŸ")
            
            else:
                st.info("æŒ‡å®šæœŸé–“å†…ã«ç´å…¥é€²åº¦ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“")
        
        except Exception as e:
            st.error(f"é€²åº¦ä¸€è¦§ã‚¨ãƒ©ãƒ¼: {e}")
    
    def _show_matrix_view(self, progress_df: pd.DataFrame, can_edit):
        """ãƒãƒˆãƒªãƒƒã‚¯ã‚¹è¡¨ç¤ºï¼ˆæ¨ªè»¸=æ—¥ä»˜ã€ç¸¦è»¸=è£½å“ã‚³ãƒ¼ãƒ‰Ã—çŠ¶æ…‹ï¼‰- ç·¨é›†å¯èƒ½"""
        
        # è£½å“åãƒãƒƒãƒ”ãƒ³ã‚°ä½œæˆ
        product_names = progress_df.groupby('product_code')['product_name'].first().to_dict()
        
        # è£½å“ã‚³ãƒ¼ãƒ‰ä¸€è¦§ã‚’å–å¾—
        product_codes = sorted(progress_df['product_code'].unique())
        
        # æ—¥ä»˜ä¸€è¦§ã‚’å–å¾—ï¼ˆæ–‡å­—åˆ—å½¢å¼ï¼‰
        dates = sorted(progress_df['delivery_date'].unique())
        date_columns = [d.strftime('%mæœˆ%dæ—¥') for d in dates]
        
        st.write(f"**è£½å“æ•°**: {len(product_codes)}")
        st.write(f"**æ—¥ä»˜æ•°**: {len(dates)}")
        
        # ã‚ªãƒ¼ãƒ€ãƒ¼IDãƒãƒƒãƒ”ãƒ³ã‚°ï¼ˆæ›´æ–°ç”¨ï¼‰
        order_mapping = {}  # {(product_code, date_str): order_id}
        for _, row in progress_df.iterrows():
            key = (row['product_code'], row['delivery_date'].strftime('%mæœˆ%dæ—¥'))
            order_mapping[key] = row['id']
        
        # çµæœã‚’æ ¼ç´ã™ã‚‹ãƒªã‚¹ãƒˆ
        result_rows = []
        
        for product_code in product_codes:
            product_data = progress_df[progress_df['product_code'] == product_code]
            
            # å„æŒ‡æ¨™ã®è¡Œã‚’ä½œæˆ
            order_row = {'è£½å“ã‚³ãƒ¼ãƒ‰': product_code, 'çŠ¶æ…‹': 'å—æ³¨æ•°', 'row_type': 'order'}
            planned_row = {'è£½å“ã‚³ãƒ¼ãƒ‰': '', 'çŠ¶æ…‹': 'ç´å…¥è¨ˆç”»æ•°', 'row_type': 'planned'}
            planned_progress_row = {'è£½å“ã‚³ãƒ¼ãƒ‰': '', 'çŠ¶æ…‹': 'è¨ˆç”»é€²åº¦', 'row_type': 'planned_progress'}
            shipped_row = {'è£½å“ã‚³ãƒ¼ãƒ‰': '', 'çŠ¶æ…‹': 'ç´å…¥å®Ÿç¸¾', 'row_type': 'shipped'}
            progress_row = {'è£½å“ã‚³ãƒ¼ãƒ‰': '', 'çŠ¶æ…‹': 'é€²åº¦', 'row_type': 'progress'}
            keisen_row = {'è£½å“ã‚³ãƒ¼ãƒ‰': '', 'çŠ¶æ…‹': '___', 'row_type': 'ãƒ¼ãƒ¼ãƒ¼'}
            
            cumulative_order = 0
            cumulative_planned = 0
            cumulative_shipped = 0
            
            for idx, (date_obj, date_str) in enumerate(zip(dates, date_columns)):
                # ãã®æ—¥ã®ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                day_data = product_data[product_data['delivery_date'] == date_obj]
                
                if not day_data.empty:
                    row = day_data.iloc[0]
                    
                    order_qty = int(row['order_quantity']) if pd.notna(row['order_quantity']) else 0
                    
                    # planned_quantity ã®å®‰å…¨ãªå–å¾—
                    if 'planned_quantity' in day_data.columns and pd.notna(row['planned_quantity']):
                        planned_qty = int(row['planned_quantity'])
                    else:
                        planned_qty = 0
                    
                    # shipped_quantity ã®å®‰å…¨ãªå–å¾—
                    if 'shipped_quantity' in day_data.columns and pd.notna(row['shipped_quantity']):
                        shipped_qty = int(row['shipped_quantity'])
                    else:
                        shipped_qty = 0
                    
                    cumulative_order += order_qty
                    cumulative_planned += planned_qty
                    cumulative_shipped += shipped_qty
                    
                    order_row[date_str] = order_qty
                    planned_row[date_str] = planned_qty
                    planned_progress_row[date_str] = cumulative_planned - cumulative_order
                    shipped_row[date_str] = shipped_qty
                else:
                    order_row[date_str] = 0
                    planned_row[date_str] = 0
                    planned_progress_row[date_str] = cumulative_planned - cumulative_order
                    shipped_row[date_str] = 0
                
                # é€²åº¦ = ç´¯è¨ˆå‡ºè· - ç´¯è¨ˆå—æ³¨
                progress = cumulative_shipped - cumulative_order
                progress_row[date_str] = int(progress)
            
            result_rows.extend([order_row, planned_row, planned_progress_row, shipped_row, progress_row, keisen_row])
        
        # DataFrameã«å¤‰æ›
        if not result_rows:
            st.info("è¡¨ç¤ºã™ã‚‹ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“")
            return
        
        result_df = pd.DataFrame(result_rows)
        
        # ã‚«ãƒ©ãƒ ã®é †åºã‚’æ•´ç†ï¼ˆå­˜åœ¨ã™ã‚‹ã‚«ãƒ©ãƒ ã®ã¿ï¼‰
        columns = ['è£½å“ã‚³ãƒ¼ãƒ‰', 'çŠ¶æ…‹', 'row_type'] + date_columns
        # å­˜åœ¨ã—ãªã„ã‚«ãƒ©ãƒ ã‚’é™¤å¤–
        columns = [col for col in columns if col in result_df.columns]
        result_df = result_df[columns]
        
        st.write("---")
        st.write("**æ—¥ä»˜Ã—è£½å“ãƒãƒˆãƒªãƒƒã‚¯ã‚¹ï¼ˆå—æ³¨ãƒ»è¨ˆç”»ãƒ»å®Ÿç¸¾ãƒ»é€²åº¦ï¼‰**")
        
        # ä¿®æ­£: åˆ—ã‚’å›ºå®šè¡¨ç¤ºï¼ˆè£½å“ã‚³ãƒ¼ãƒ‰ã¨çŠ¶æ…‹åˆ—ã‚’å›ºå®šï¼‰
        edited_df = st.data_editor(
            result_df,
            use_container_width=True,
            hide_index=True,
            disabled=['è£½å“ã‚³ãƒ¼ãƒ‰', 'çŠ¶æ…‹', 'row_type'],  # ç·¨é›†ä¸å¯ã‚«ãƒ©ãƒ 
            column_config={
                "è£½å“ã‚³ãƒ¼ãƒ‰": st.column_config.TextColumn(
                    "è£½å“ã‚³ãƒ¼ãƒ‰", 
                    width="medium",
                    pinned=True
                ),
                "çŠ¶æ…‹": st.column_config.TextColumn(
                    "çŠ¶æ…‹", 
                    width="small",
                    pinned=True
                ),
                "row_type": None,  # éè¡¨ç¤º
                **{col: st.column_config.NumberColumn(col, step=1) for col in date_columns}
            },
            key="matrix_editor"
        )
        
        # ä¿å­˜ãƒœã‚¿ãƒ³
        col_save1, col_save2 = st.columns([1, 5])
        
        with col_save1:
            if st.button("ğŸ’¾ å¤‰æ›´ã‚’ä¿å­˜", type="primary", use_container_width=True, disabled=not can_edit):
                # å¤‰æ›´ã‚’æ¤œå‡ºã—ã¦ä¿å­˜
                changes_saved = self._save_matrix_changes(
                    original_df=result_df,
                    edited_df=edited_df,
                    order_mapping=order_mapping,
                    product_codes=product_codes,
                    dates=dates,
                    date_columns=date_columns,
                    progress_df=progress_df
                )
                
                if changes_saved:
                    st.success("âœ… å¤‰æ›´ã‚’ä¿å­˜ã—ã¾ã—ãŸ")
                    st.rerun()
                else:
                    st.info("å¤‰æ›´ã¯ã‚ã‚Šã¾ã›ã‚“ã§ã—ãŸ")
        
        with col_save2:
            st.caption("â€» ã€Œè¨ˆç”»é€²åº¦ã€ã€Œé€²åº¦ã€è¡Œã¯è‡ªå‹•è¨ˆç®—ã•ã‚Œã¾ã™ï¼ˆè¨ˆç”»é€²åº¦=ç´¯è¨ˆè¨ˆç”» - ç´¯è¨ˆå—æ³¨ã€é€²åº¦=ç´¯è¨ˆå‡ºè· - ç´¯è¨ˆå—æ³¨ï¼‰")
        
        # èª¬æ˜
        with st.expander("ğŸ“‹ è¡¨ã®è¦‹æ–¹"):
            st.write("""
            **å„è¡Œã®æ„å‘³:**
            - **å—æ³¨æ•°**: ãã®æ—¥ã®å—æ³¨æ•°é‡ï¼ˆç·¨é›†ä¸å¯ï¼‰
            - **ç´å…¥è¨ˆç”»æ•°**: ç©è¼‰è¨ˆç”»ã§è¨­å®šã•ã‚ŒãŸæ•°é‡ï¼ˆç·¨é›†å¯ï¼‰
            - **è¨ˆç”»é€²åº¦**: ç´¯è¨ˆè¨ˆç”» - ç´¯è¨ˆå—æ³¨ï¼ˆè‡ªå‹•è¨ˆç®—ï¼‰
            - **ç´å…¥å®Ÿç¸¾**: å®Ÿéš›ã«å‡ºè·ã—ãŸæ•°é‡ï¼ˆç·¨é›†å¯ï¼‰
            - **é€²åº¦**: ç´¯è¨ˆå‡ºè· - ç´¯è¨ˆå—æ³¨ï¼ˆè‡ªå‹•è¨ˆç®—ã€ãƒã‚¤ãƒŠã‚¹ã¯æœªç´åˆ†ï¼‰
            
            **ç·¨é›†æ–¹æ³•:**
            1. ã€Œç´å…¥è¨ˆç”»æ•°ã€ã¾ãŸã¯ã€Œç´å…¥å®Ÿç¸¾ã€ã®ã‚»ãƒ«ã‚’ãƒ€ãƒ–ãƒ«ã‚¯ãƒªãƒƒã‚¯
            2. æ•°å€¤ã‚’å…¥åŠ›
            3. ã€ŒğŸ’¾ å¤‰æ›´ã‚’ä¿å­˜ã€ãƒœã‚¿ãƒ³ã‚’ã‚¯ãƒªãƒƒã‚¯
            """)

    def _save_matrix_changes(self, original_df, edited_df, order_mapping, 
                            product_codes, dates, date_columns, progress_df):
        """ãƒãƒˆãƒªãƒƒã‚¯ã‚¹ã®å¤‰æ›´ã‚’ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ã«ä¿å­˜"""
        
        changes_made = False
        
        for product_code in product_codes:
            for date_obj, date_str in zip(dates, date_columns):
                # ã‚ªãƒ¼ãƒ€ãƒ¼IDã‚’å–å¾—
                order_key = (product_code, date_str)
                if order_key not in order_mapping:
                    continue
                
                order_id = order_mapping[order_key]
                
                # å…ƒãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                original_data = progress_df[
                    (progress_df['product_code'] == product_code) & 
                    (progress_df['delivery_date'] == date_obj)
                ]
                
                if original_data.empty:
                    continue
                
                # NaNå¯¾å¿œ
                original_planned = int(original_data['planned_quantity'].iloc[0]) if pd.notna(original_data['planned_quantity'].iloc[0]) else 0
                original_shipped = int(original_data['shipped_quantity'].iloc[0]) if pd.notna(original_data['shipped_quantity'].iloc[0]) else 0
                
                # ç·¨é›†å¾Œã®ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
                planned_rows = edited_df[
                    (edited_df['row_type'] == 'planned') &
                    ((edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == product_code) | (edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == ''))
                ]
                
                shipped_rows = edited_df[
                    (edited_df['row_type'] == 'shipped') &
                    ((edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == product_code) | (edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == ''))
                ]
                
                # ç´å…¥è¨ˆç”»æ•°ã®å¤‰æ›´ãƒã‚§ãƒƒã‚¯
                if not planned_rows.empty and date_str in planned_rows.columns:
                    product_planned_rows = planned_rows[
                        (planned_rows.index > edited_df[edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == product_code].index.min()) &
                        (planned_rows.index < edited_df[edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == product_code].index.min() + 4)
                    ]
                    
                    if not product_planned_rows.empty:
                        new_planned = int(product_planned_rows.iloc[0][date_str]) if pd.notna(product_planned_rows.iloc[0][date_str]) else 0
                        
                        if new_planned != original_planned:
                            update_data = {'planned_quantity': new_planned}
                            success = self.service.update_delivery_progress(order_id, update_data)
                            if success:
                                changes_made = True
                                print(f"âœ… è¨ˆç”»æ•°æ›´æ–°: order_id={order_id}, {original_planned} â†’ {new_planned}")
                
                # ç´å…¥å®Ÿç¸¾ã®å¤‰æ›´ãƒã‚§ãƒƒã‚¯
                if not shipped_rows.empty and date_str in shipped_rows.columns:
                    product_shipped_rows = shipped_rows[
                        (shipped_rows.index > edited_df[edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == product_code].index.min()) &
                        (shipped_rows.index < edited_df[edited_df['è£½å“ã‚³ãƒ¼ãƒ‰'] == product_code].index.min() + 4)
                    ]
                    
                    if not product_shipped_rows.empty:
                        new_shipped = int(product_shipped_rows.iloc[0][date_str]) if pd.notna(product_shipped_rows.iloc[0][date_str]) else 0
                        
                        # âœ… ä¿®æ­£: ç›´æ¥ delivery_progress ã‚’æ›´æ–°
                        if new_shipped != original_shipped:
                            # 1. delivery_progress.shipped_quantity ã‚’ç›´æ¥æ›´æ–°
                            update_data = {'shipped_quantity': new_shipped}
                            success = self.service.update_delivery_progress(order_id, update_data)
                            
                            if success:
                                changes_made = True
                                print(f"âœ… å®Ÿç¸¾æ›´æ–°: order_id={order_id}, {original_shipped} â†’ {new_shipped}")
                                
                                # 2. å·®åˆ†ãŒã‚ã‚Œã°å‡ºè·å®Ÿç¸¾ãƒ¬ã‚³ãƒ¼ãƒ‰ã‚‚ä½œæˆï¼ˆå±¥æ­´ã¨ã—ã¦ï¼‰
                                diff = new_shipped - original_shipped
                                if diff > 0:
                                    shipment_data = {
                                        'progress_id': order_id,
                                        'truck_id': 1,
                                        'shipment_date': date_obj,
                                        'shipped_quantity': diff,
                                        'driver_name': 'ãƒãƒˆãƒªãƒƒã‚¯ã‚¹å…¥åŠ›',
                                        'actual_departure_time': None,
                                        'actual_arrival_time': None,
                                        'notes': f'ãƒãƒˆãƒªãƒƒã‚¯ã‚¹ã‹ã‚‰ç›´æ¥å…¥åŠ›ï¼ˆç´¯è¨ˆ: {new_shipped}ï¼‰'
                                    }
                                    self.service.create_shipment_record(shipment_data)
        
        return changes_made

    def _show_progress_registration(self, can_edit):
        """æ–°è¦ç™»éŒ²"""
        st.header("â• æ–°è¦ç´å…¥é€²åº¦ç™»éŒ²")

        if not can_edit:
            st.info("ç·¨é›†æ¨©é™ãŒãªã„ãŸã‚ã€æ–°è¦ç™»éŒ²ã¯ã§ãã¾ã›ã‚“")
            return
        
        with st.form("create_progress_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**ã‚ªãƒ¼ãƒ€ãƒ¼æƒ…å ±**")
                order_id = st.text_input("ã‚ªãƒ¼ãƒ€ãƒ¼ID *", placeholder="ä¾‹: ORD-2025-001")
                
                # è£½å“é¸æŠ
                try:
                    products = self.service.product_repo.get_all_products()
                    if not products.empty:
                        product_options = {
                            f"{row['product_code']} - {row['product_name']}": row['id']
                            for _, row in products.iterrows()
                        }
                        selected_product = st.selectbox("è£½å“ *", options=list(product_options.keys()))
                        product_id = product_options[selected_product]
                    else:
                        st.warning("è£½å“ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
                        product_id = None
                except:
                    st.error("è£½å“ãƒ‡ãƒ¼ã‚¿å–å¾—ã‚¨ãƒ©ãƒ¼")
                    product_id = None
                
                order_date = st.date_input("å—æ³¨æ—¥ *", value=date.today())
                delivery_date = st.date_input("ç´æœŸ *", value=date.today() + timedelta(days=7))
                order_quantity = st.number_input("å—æ³¨æ•°é‡ *", min_value=1, value=100, step=1)
            
            with col2:
                st.write("**å¾—æ„å…ˆæƒ…å ±**")
                customer_code = st.text_input("å¾—æ„å…ˆã‚³ãƒ¼ãƒ‰", placeholder="ä¾‹: C001")
                customer_name = st.text_input("å¾—æ„å…ˆå", placeholder="ä¾‹: æ ªå¼ä¼šç¤¾ã€‡ã€‡")
                delivery_location = st.text_input("ç´å…¥å…ˆ", placeholder="ä¾‹: æ±äº¬å·¥å ´")
                priority = st.number_input("å„ªå…ˆåº¦ï¼ˆ1-10ï¼‰", min_value=1, max_value=10, value=5)
                notes = st.text_area("å‚™è€ƒ")
            
            submitted = st.form_submit_button("â• ç™»éŒ²", type="primary")
            
            if submitted:
                if not order_id or not product_id:
                    st.error("ã‚ªãƒ¼ãƒ€ãƒ¼IDã¨è£½å“ã¯å¿…é ˆã§ã™")
                else:
                    progress_data = {
                        'order_id': order_id,
                        'product_id': product_id,
                        'order_date': order_date,
                        'delivery_date': delivery_date,
                        'order_quantity': order_quantity,
                        'customer_code': customer_code,
                        'customer_name': customer_name,
                        'delivery_location': delivery_location,
                        'priority': priority,
                        'notes': notes
                    }
                    
                    progress_id = self.service.create_delivery_progress(progress_data)
                    if progress_id > 0:
                        st.success(f"ç´å…¥é€²åº¦ã‚’ç™»éŒ²ã—ã¾ã—ãŸï¼ˆID: {progress_id}ï¼‰")
                        st.balloons()
                        st.rerun()
                    else:
                        st.error("ç´å…¥é€²åº¦ç™»éŒ²ã«å¤±æ•—ã—ã¾ã—ãŸ")
    
    def _show_actual_registration(self, can_edit):
        """å®Ÿç¸¾ç™»éŒ²"""
        st.header("âœ… ç©è¾¼å®Ÿç¸¾ç™»éŒ²")

        if not can_edit:
            st.info("ç·¨é›†æ¨©é™ãŒãªã„ãŸã‚ã€å®Ÿç¸¾ç™»éŒ²ã¯ã§ãã¾ã›ã‚“")
            return
        
        try:
            trucks_df = self.service.get_trucks()
        except Exception as e:
            st.error(f"ãƒˆãƒ©ãƒƒã‚¯æƒ…å ±ã®å–å¾—ã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
            return
        
        if trucks_df is None or trucks_df.empty:
            st.info("ãƒˆãƒ©ãƒƒã‚¯ãƒã‚¹ã‚¿ãŒç©ºã§ã™ã€‚å…ˆã«ãƒˆãƒ©ãƒƒã‚¯ã‚’ç™»éŒ²ã—ã¦ãã ã•ã„ã€‚")
            return
        
        truck_options = {
            str(row["name"]): int(row["id"])
            for _, row in trucks_df.iterrows()
            if pd.notna(row.get("name")) and pd.notna(row.get("id"))
        }
        
        if not truck_options:
            st.info("é¸æŠå¯èƒ½ãªãƒˆãƒ©ãƒƒã‚¯ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            return
        
        col1, col2 = st.columns(2)
        with col1:
            loading_date = st.date_input(
                "ç©è¾¼æ—¥",
                value=date.today(),
                key="actual_loading_date"
            )
        with col2:
            truck_name = st.selectbox(
                "ãƒˆãƒ©ãƒƒã‚¯",
                options=list(truck_options.keys()),
                key="actual_truck_select"
            )
        
        selected_truck_id = truck_options.get(truck_name)
        if not selected_truck_id:
            st.warning("ãƒˆãƒ©ãƒƒã‚¯ã‚’é¸æŠã—ã¦ãã ã•ã„ã€‚")
            return
        
        try:
            plan_items = self.service.get_loading_plan_details_by_date(loading_date, selected_truck_id)
        except Exception as e:
            st.error(f"ç©è¼‰è¨ˆç”»ã®å–å¾—ã«å¤±æ•—ã—ã¾ã—ãŸ: {e}")
            return
        
        if not plan_items:
            st.info("æŒ‡å®šæ¡ä»¶ã«è©²å½“ã™ã‚‹ç©è¼‰è¨ˆç”»ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            return
        
        plan_df = pd.DataFrame(plan_items)
        if plan_df.empty or 'id' not in plan_df.columns:
            st.error("ç©è¼‰è¨ˆç”»æ˜ç´°ã®å½¢å¼ãŒä¸æ­£ã§ã™ã€‚")
            return
        
        plan_df = plan_df.set_index('id')
        
        if 'delivery_date' in plan_df.columns:
            plan_df['delivery_date'] = pd.to_datetime(
                plan_df['delivery_date'], errors='coerce'
            ).dt.date
        plan_df['delivery_date'] = plan_df['delivery_date'].fillna(loading_date)
        
        if 'trip_number' in plan_df.columns:
            plan_df['trip_number'] = pd.to_numeric(plan_df['trip_number'], errors='coerce').fillna(1).astype(int)
        else:
            plan_df['trip_number'] = 1
        
        plan_df['num_containers'] = pd.to_numeric(plan_df.get('num_containers', 0), errors='coerce').fillna(0).astype(int)
        plan_df['total_quantity'] = pd.to_numeric(plan_df.get('total_quantity', 0), errors='coerce').fillna(0).astype(int)
        plan_df['planned_quantity'] = plan_df['total_quantity']
        
        progress_cache: Dict[int, Optional[Dict[str, Any]]] = {}
        missing_progress: list[str] = []
        
        plan_df['current_shipped'] = None
        plan_df['current_status'] = None
        
        for detail_id, row in plan_df.iterrows():
            product_id = row.get('product_id')
            try:
                product_id_int = int(product_id)
            except (TypeError, ValueError):
                progress_cache[detail_id] = None
                missing_progress.append(f"{row.get('product_code', '') or 'ä¸æ˜'}")
                continue
            
            delivery_value = row.get('delivery_date') or loading_date
            if isinstance(delivery_value, pd.Timestamp):
                delivery_value = delivery_value.to_pydatetime().date()
            elif isinstance(delivery_value, datetime):
                delivery_value = delivery_value.date()
            elif isinstance(delivery_value, str):
                try:
                    delivery_value = datetime.strptime(delivery_value, "%Y-%m-%d").date()
                except ValueError:
                    delivery_value = loading_date
            
            plan_df.at[detail_id, 'delivery_date'] = delivery_value
            
            try:
                progress = self.service.get_delivery_progress_by_product_and_date(product_id_int, delivery_value)
            except Exception as e:
                st.warning(f"ç´å…¥é€²åº¦ã®å–å¾—ã«å¤±æ•—ã—ã¾ã—ãŸï¼ˆè£½å“ID:{product_id_int}ï¼‰: {e}")
                progress = None
            
            progress_cache[detail_id] = progress
            
            if progress:
                shipped_val = progress.get('shipped_quantity')
                plan_df.at[detail_id, 'current_shipped'] = int(shipped_val) if shipped_val is not None else 0
                plan_df.at[detail_id, 'current_status'] = progress.get('status')
            else:
                plan_df.at[detail_id, 'current_shipped'] = None
                plan_df.at[detail_id, 'current_status'] = None
                missing_progress.append(f"{row.get('product_code', '') or 'ä¸æ˜'}ï¼ˆ{delivery_value}ï¼‰")
        
        product_codes = plan_df.get('product_code', pd.Series('', index=plan_df.index))
        product_names = plan_df.get('product_name', pd.Series('', index=plan_df.index))
        
        display_df = pd.DataFrame(
            {
                "ç©è¾¼é †": plan_df['trip_number'],
                "è£½å“ã‚³ãƒ¼ãƒ‰": product_codes,
                "è£½å“å": product_names,
                "ç´å…¥æ—¥": plan_df['delivery_date'],
                "è¨ˆç”»æ•°é‡": plan_df['planned_quantity'],
                "æ—¢å‡ºè·æ•°é‡": plan_df['current_shipped'].fillna(0).astype(int),
                "å®Ÿç¸¾æ•°é‡": plan_df['planned_quantity']
            },
            index=plan_df.index
        )
        display_df.index.name = "detail_id"
        
        st.caption("è¨ˆç”»æ•°é‡ã‚’ãƒ™ãƒ¼ã‚¹ã«å®Ÿç¸¾æ•°é‡ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„ã€‚ä¸è¦ãªè¡Œã¯0ã®ã¾ã¾ã«ã—ã¾ã™ã€‚")
        if missing_progress:
            st.warning("ç´å…¥é€²åº¦ãŒè¦‹ã¤ã‹ã‚‰ãªã„æ˜ç´°ãŒã‚ã‚Šã¾ã™: " + "ã€".join(sorted(set(missing_progress))))
        
        form_key = f"actual_registration_form_{selected_truck_id}_{loading_date.isoformat()}"
        with st.form(form_key):
            edited_df = st.data_editor(
                display_df,
                key=f"actual_editor_{selected_truck_id}_{loading_date.isoformat()}",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "ç©è¾¼é †": st.column_config.NumberColumn("ç©è¾¼é †", disabled=True),
                    "è£½å“ã‚³ãƒ¼ãƒ‰": st.column_config.TextColumn("è£½å“ã‚³ãƒ¼ãƒ‰", disabled=True),
                    "è£½å“å": st.column_config.TextColumn("è£½å“å", disabled=True),
                    "ç´å…¥æ—¥": st.column_config.DateColumn("ç´å…¥æ—¥", disabled=True, format="YYYY-MM-DD"),
                    "è¨ˆç”»æ•°é‡": st.column_config.NumberColumn("è¨ˆç”»æ•°é‡", disabled=True),
                    "æ—¢å‡ºè·æ•°é‡": st.column_config.NumberColumn("æ—¢å‡ºè·æ•°é‡", disabled=True),
                    "å®Ÿç¸¾æ•°é‡": st.column_config.NumberColumn("å®Ÿç¸¾æ•°é‡", min_value=0, step=1)
                }
            )
            
            driver_name = st.text_input("ãƒ‰ãƒ©ã‚¤ãƒãƒ¼å", key=f"actual_driver_{selected_truck_id}")
            notes = st.text_area(
                "å‚™è€ƒï¼ˆå¿…è¦ã«å¿œã˜ã¦å…¥åŠ›ï¼‰",
                key=f"actual_notes_{selected_truck_id}",
                placeholder=f"ä¾‹: {truck_name} {loading_date} ç©è¾¼"
            )
            
            submitted = st.form_submit_button("å®Ÿç¸¾ã‚’ç™»éŒ²", type="primary")
            
            if submitted:
                if edited_df.empty:
                    st.info("ç™»éŒ²å¯¾è±¡ã®æ˜ç´°ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
                    return
                
                registered = 0
                failed_entries: list[str] = []
                missing_entries: list[str] = []
                
                for detail_id, row in edited_df.iterrows():
                    try:
                        detail_id_int = int(detail_id)
                    except (TypeError, ValueError):
                        continue
                    
                    quantity_value = pd.to_numeric(row.get("å®Ÿç¸¾æ•°é‡"), errors='coerce')
                    if pd.isna(quantity_value) or quantity_value <= 0:
                        continue
                    
                    progress = progress_cache.get(detail_id_int)
                    plan_row = plan_df.loc[detail_id_int]
                    
                    if not progress:
                        missing_entries.append(f"{plan_row.get('product_code', '') or 'ä¸æ˜'}ï¼ˆ{plan_row.get('delivery_date')}ï¼‰")
                        continue
                    
                    shipment_data = {
                        'progress_id': progress['id'],
                        'truck_id': selected_truck_id,
                        'shipment_date': loading_date,
                        'shipped_quantity': int(quantity_value),
                        'container_id': plan_row.get('container_id'),
                        'num_containers': plan_row.get('num_containers'),
                        'driver_name': driver_name,
                        'notes': notes
                    }
                    
                    success = self.service.create_shipment_record(shipment_data)
                    if success:
                        registered += 1
                    else:
                        failed_entries.append(f"{plan_row.get('product_code', '') or 'ä¸æ˜'}ï¼ˆ{plan_row.get('delivery_date')}ï¼‰")
                
                if registered:
                    st.success(f"{registered} ä»¶ã®å®Ÿç¸¾ã‚’ç™»éŒ²ã—ã¾ã—ãŸã€‚")
                    st.balloons()
                if failed_entries:
                    st.error("ç™»éŒ²ã«å¤±æ•—ã—ãŸæ˜ç´°: " + "ã€".join(failed_entries))
                if missing_entries:
                    st.warning("ç´å…¥é€²åº¦ãŒè¦‹ã¤ã‹ã‚‰ãšç™»éŒ²ã§ããªã‹ã£ãŸæ˜ç´°: " + "ã€".join(missing_entries))
                
                if registered and not failed_entries:
                    st.info("ä»–ã®ã‚¿ãƒ–ã§æœ€æ–°ã®å®Ÿç¸¾ã‚’ç¢ºèªã§ãã¾ã™ã€‚")
                    st.rerun()

        # âœ… å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—ã‚»ã‚¯ã‚·ãƒ§ãƒ³
        st.markdown("---")
        with st.expander("ğŸ”„ å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—"):
            st.write("å‡ºè·å®Ÿç¸¾ç™»éŒ²å¾Œã€ç´å…¥é€²åº¦ã®å‡ºè·æ¸ˆæ•°é‡ï¼ˆshipped_remaining_quantityï¼‰ã‚’å†è¨ˆç®—ã—ã¾ã™ã€‚")

            # è£½å“ãƒªã‚¹ãƒˆã‚’å–å¾—
            try:
                products = self.service.product_repo.get_all_products()
                if not products.empty:
                    sr_product_options = {
                        f"{row['product_code']} - {row['product_name']}": row['id']
                        for _, row in products.iterrows()
                    }
                    sr_selected_product = st.selectbox(
                        "è£½å“ã‚³ãƒ¼ãƒ‰",
                        options=list(sr_product_options.keys()),
                        key="actual_reg_sr_product_select"
                    )
                    sr_product_id = sr_product_options[sr_selected_product]
                else:
                    st.warning("è£½å“ãŒç™»éŒ²ã•ã‚Œã¦ã„ã¾ã›ã‚“")
                    sr_product_id = None
            except Exception as e:
                st.error(f"è£½å“ãƒ‡ãƒ¼ã‚¿å–å¾—ã‚¨ãƒ©ãƒ¼: {e}")
                sr_product_id = None

            sr_start_date = st.date_input("å†è¨ˆç®—é–‹å§‹æ—¥ï¼ˆå®Ÿç¸¾ï¼‰", key="actual_reg_sr_start_date")
            sr_end_date = st.date_input("å†è¨ˆç®—çµ‚äº†æ—¥ï¼ˆå®Ÿç¸¾ï¼‰", key="actual_reg_sr_end_date")

            col_sr_one, col_sr_all = st.columns(2)

            with col_sr_one:
                if st.button("é¸æŠè£½å“ã®å®Ÿç¸¾é€²åº¦ã‚’å†è¨ˆç®—", key="actual_reg_btn_sr_one", disabled=not can_edit):
                    if sr_product_id:
                        self.service.recompute_shipped_remaining(sr_product_id, sr_start_date, sr_end_date)
                        st.success("å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—ãŒå®Œäº†ã—ã¾ã—ãŸ")
                    else:
                        st.error("è£½å“ã‚’é¸æŠã—ã¦ãã ã•ã„")

            with col_sr_all:
                if st.button("å…¨è£½å“ã®å®Ÿç¸¾é€²åº¦ã‚’å†è¨ˆç®—", key="actual_reg_btn_sr_all", disabled=not can_edit):
                    self.service.recompute_shipped_remaining_all(sr_start_date, sr_end_date)
                    st.success("å…¨è£½å“ã®å®Ÿç¸¾é€²åº¦ã®å†è¨ˆç®—ãŒå®Œäº†ã—ã¾ã—ãŸ")

    def _show_shipment_records(self):
        """å‡ºè·å®Ÿç¸¾è¡¨ç¤º"""
        st.header("ğŸ“¦ å‡ºè·å®Ÿç¸¾ä¸€è¦§")
        
        # ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼
        col_f1, col_f2 = st.columns(2)
        
        with col_f1:
            filter_start = st.date_input(
                "å‡ºè·æ—¥ï¼ˆé–‹å§‹ï¼‰",
                value=date.today() - timedelta(days=7),
                key="shipment_start_filter"
            )
        
        with col_f2:
            filter_end = st.date_input(
                "å‡ºè·æ—¥ï¼ˆçµ‚äº†ï¼‰",
                value=date.today(),
                key="shipment_end_filter"
            )
        
        try:
            shipment_df = self.service.get_shipment_records()
            
            if not shipment_df.empty:
                # æ—¥ä»˜ãƒ•ã‚£ãƒ«ã‚¿ãƒ¼é©ç”¨
                shipment_df['shipment_date'] = pd.to_datetime(shipment_df['shipment_date']).dt.date
                filtered_df = shipment_df[
                    (shipment_df['shipment_date'] >= filter_start) &
                    (shipment_df['shipment_date'] <= filter_end)
                ]
                
                if not filtered_df.empty:
                    # è¡¨ç¤ºç”¨ãƒ‡ãƒ¼ã‚¿ãƒ•ãƒ¬ãƒ¼ãƒ 
                    display_cols = ['shipment_date', 'order_id', 'product_code', 'product_name', 
                                  'truck_name', 'shipped_quantity', 'driver_name']
                    
                    # ã‚«ãƒ©ãƒ ãŒå­˜åœ¨ã™ã‚‹ã‹ãƒã‚§ãƒƒã‚¯
                    available_cols = [col for col in display_cols if col in filtered_df.columns]
                    
                    if 'num_containers' in filtered_df.columns:
                        available_cols.append('num_containers')
                    
                    display_df = filtered_df[available_cols].copy()
                    
                    # ã‚«ãƒ©ãƒ åã‚’æ—¥æœ¬èªã«
                    column_mapping = {
                        'shipment_date': 'å‡ºè·æ—¥',
                        'order_id': 'ã‚ªãƒ¼ãƒ€ãƒ¼ID',
                        'product_code': 'è£½å“ã‚³ãƒ¼ãƒ‰',
                        'product_name': 'è£½å“å',
                        'truck_name': 'ãƒˆãƒ©ãƒƒã‚¯',
                        'shipped_quantity': 'å‡ºè·æ•°é‡',
                        'num_containers': 'å®¹å™¨æ•°',
                        'driver_name': 'ãƒ‰ãƒ©ã‚¤ãƒãƒ¼'
                    }
                    
                    display_df.columns = [column_mapping.get(col, col) for col in display_df.columns]
                    
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "å‡ºè·æ—¥": st.column_config.DateColumn("å‡ºè·æ—¥", format="YYYY-MM-DD"),
                        }
                    )
                    
                    # çµ±è¨ˆæƒ…å ±
                    st.subheader("ğŸ“Š å‡ºè·çµ±è¨ˆ")
                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    
                    with col_stat1:
                        total_shipments = len(filtered_df)
                        st.metric("ç·å‡ºè·å›æ•°", f"{total_shipments}å›")
                    
                    with col_stat2:
                        total_quantity = filtered_df['shipped_quantity'].sum()
                        st.metric("ç·å‡ºè·æ•°é‡", f"{total_quantity:,.0f}å€‹")
                    
                    with col_stat3:
                        unique_products = filtered_df['product_id'].nunique()
                        st.metric("å‡ºè·è£½å“ç¨®é¡", f"{unique_products}ç¨®")
                else:
                    st.info("æŒ‡å®šæœŸé–“å†…ã®å‡ºè·å®Ÿç¸¾ãŒã‚ã‚Šã¾ã›ã‚“")
            else:
                st.info("å‡ºè·å®Ÿç¸¾ãŒã‚ã‚Šã¾ã›ã‚“")
        
        except Exception as e:
            st.error(f"å‡ºè·å®Ÿç¸¾å–å¾—ã‚¨ãƒ©ãƒ¼: {e}")

    def _show_internal_orders(self):
        """ç¤¾å†…æ³¨æ–‡ï¼ˆè£½é€ å·¥ç¨‹ï¼‰ã‚¿ãƒ–"""
        st.header("ğŸ­ ç¤¾å†…æ³¨æ–‡")
        st.write("ç©è¼‰è¨ˆç”»ã§è¨­å®šã•ã‚ŒãŸæ•°é‡ã‚’ç¤¾å†…å‘ã‘ã®åŠ å·¥æŒ‡ç¤ºã¨ã—ã¦ç¢ºèªã§ãã¾ã™ã€‚")

        st.markdown("---")
        st.subheader("ğŸ“… è¡¨ç¤ºæœŸé–“ï¼ˆç©è¼‰æ—¥åŸºæº–ï¼‰")
        col1, col2 = st.columns(2)

        with col1:
            start_date = st.date_input(
                "é–‹å§‹æ—¥ï¼ˆç©è¼‰æ—¥ï¼‰",
                value=date.today(),
                key="internal_order_start_date"
            )

        with col2:
            end_date = st.date_input(
                "çµ‚äº†æ—¥ï¼ˆç©è¼‰æ—¥ï¼‰",
                value=date.today() + timedelta(days=7),
                key="internal_order_end_date"
            )

        if start_date > end_date:
            st.error("é–‹å§‹æ—¥ã¯çµ‚äº†æ—¥ä»¥å‰ã®æ—¥ä»˜ã‚’é¸æŠã—ã¦ãã ã•ã„ã€‚")
            return

        # delivery_progressã‹ã‚‰ç©è¼‰è¨ˆç”»æ•°ã‚’å–å¾—
        try:
            progress_df = self.service.get_delivery_progress(start_date, end_date)
        except Exception as e:
            st.error(f"ãƒ‡ãƒ¼ã‚¿å–å¾—ã‚¨ãƒ©ãƒ¼: {e}")
            return

        if progress_df is None or progress_df.empty:
            st.info("é¸æŠã—ãŸæœŸé–“ã«å¯¾è±¡ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            return

        if 'planned_quantity' not in progress_df.columns:
            st.warning("planned_quantityåˆ—ãŒå–å¾—ã§ãã¾ã›ã‚“ã§ã—ãŸã€‚")
            return

        # ç©è¼‰è¨ˆç”»æ•°ãŒè¨­å®šã•ã‚Œã¦ã„ã‚‹è£½å“ã®ã¿
        progress_df = progress_df[pd.to_numeric(progress_df['planned_quantity'], errors='coerce').fillna(0) > 0].copy()

        if progress_df.empty:
            st.info("ç©è¼‰è¨ˆç”»ã®æ•°é‡ãŒè¨­å®šã•ã‚Œã¦ã„ã‚‹è£½å“ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            return

        progress_df['delivery_date'] = pd.to_datetime(progress_df['delivery_date']).dt.date

        st.subheader("ğŸ“‹ è£½å“åˆ¥ãƒãƒˆãƒªã‚¯ã‚¹ï¼ˆç©è¼‰æ—¥Ã—è£½å“ã‚³ãƒ¼ãƒ‰ï¼‰")
        matrix_df = self._create_internal_order_matrix(progress_df, start_date, end_date)
        if matrix_df.empty:
            st.info("è¡¨ç¤ºå¯¾è±¡ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
            return
        st.dataframe(
            matrix_df,
            use_container_width=True,
            hide_index=False,
            height=600
        )

        st.markdown("---")
        st.subheader("ğŸ“¥ Excelå‡ºåŠ›")
        excel_data = self._export_internal_orders_to_excel(matrix_df, start_date, end_date)
        filename = f"ç¤¾å†…æ³¨æ–‡_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="ğŸ“¥ ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="internal_order_excel_download"
        )

    def _create_internal_order_matrix_from_loading(self, details_df: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
        """ç©è¼‰è¨ˆç”»ã‹ã‚‰è£½å“Ã—ç©è¼‰æ—¥ã®ãƒãƒˆãƒªã‚¯ã‚¹ã‚’ä½œæˆ"""
        # è£½å“ãƒã‚¹ã‚¿ã‹ã‚‰è¡¨ç¤ºé †åºã¨è£½å“åã‚’å–å¾—
        order_map = {}
        name_map = {}
        if hasattr(self.service, "product_repo"):
            try:
                master_df = self.service.product_repo.get_all_products()
            except Exception:
                master_df = pd.DataFrame()
            else:
                if isinstance(master_df, pd.DataFrame) and not master_df.empty:
                    for _, row in master_df.iterrows():
                        code = row.get('product_code')
                        if not code:
                            continue
                        display_val = row.get('display_id')
                        if pd.notna(display_val):
                            try:
                                display_val = int(display_val)
                            except (TypeError, ValueError):
                                pass
                        else:
                            display_val = None
                        order_map[code] = display_val
                        name_map[code] = row.get('product_name', '')

        # è£½å“ã‚³ãƒ¼ãƒ‰ä¸€è¦§ã‚’å–å¾—
        product_codes = details_df['product_code'].dropna().unique().tolist()

        # è¡¨ç¤ºé †ã§ã‚½ãƒ¼ãƒˆ
        def sort_key(code: str):
            display_value = order_map.get(code) if order_map else None
            if display_value is None or pd.isna(display_value):
                display_value = float('inf')
            return (display_value, code)

        product_codes.sort(key=sort_key)

        # æ—¥ä»˜ç¯„å›²ã‚’ç”Ÿæˆ
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        date_columns = [d.strftime('%Y/%m/%d') for d in date_range]
        date_values = [d.date() for d in date_range]

        # ãƒãƒˆãƒªãƒƒã‚¯ã‚¹ãƒ‡ãƒ¼ã‚¿ã‚’ä½œæˆ
        matrix_data = []
        for product_code in product_codes:
            product_data = details_df[details_df['product_code'] == product_code]
            if product_data.empty:
                continue

            # è£½å“åã‚’å–å¾—
            product_name = name_map.get(product_code)
            if not product_name:
                product_name = product_data['product_name'].iloc[0] if 'product_name' in product_data.columns else ''

            row = {
                'è£½å“ã‚³ãƒ¼ãƒ‰': product_code,
                'è£½å“å': product_name
            }

            # å„æ—¥ä»˜ã®æ•°é‡ã‚’é›†è¨ˆ
            for date_obj, date_str in zip(date_values, date_columns):
                day_data = product_data[product_data['loading_date'] == date_obj]
                if not day_data.empty:
                    # total_quantityã‚’é›†è¨ˆ
                    total_qty = pd.to_numeric(day_data['total_quantity'], errors='coerce').fillna(0).sum()
                    row[date_str] = int(total_qty) if total_qty > 0 else 0
                else:
                    row[date_str] = 0

            matrix_data.append(row)

        # DataFrameã«å¤‰æ›
        matrix_df = pd.DataFrame(matrix_data)
        if not matrix_df.empty:
            matrix_df = matrix_df.set_index('è£½å“ã‚³ãƒ¼ãƒ‰')

        return matrix_df

    def _create_internal_order_matrix(self, progress_df: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
        """è£½å“Ã—ç´æœŸã®ãƒãƒˆãƒªã‚¯ã‚¹ã‚’ä½œæˆ"""
        order_map = {}
        name_map = {}
        if hasattr(self.service, "product_repo"):
            try:
                master_df = self.service.product_repo.get_all_products()
            except Exception:
                master_df = pd.DataFrame()
            else:
                if isinstance(master_df, pd.DataFrame) and not master_df.empty:
                    temp_order = {}
                    temp_name = {}
                    for _, row in master_df.iterrows():
                        code = row.get('product_code')
                        if not code:
                            continue
                        display_val = row.get('display_id')
                        if pd.notna(display_val):
                            try:
                                display_val = int(display_val)
                            except (TypeError, ValueError):
                                pass
                        else:
                            display_val = None
                        temp_order[code] = display_val
                        temp_name[code] = row.get('product_name', '')
                    order_map = temp_order
                    name_map = temp_name

        product_codes = progress_df['product_code'].dropna().unique().tolist()

        def sort_key(code: str):
            display_value = order_map.get(code) if order_map else None
            if display_value is None or pd.isna(display_value):
                display_value = float('inf')
            return (display_value, code)

        product_codes.sort(key=sort_key)
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        date_columns = [d.strftime('%Y/%m/%d') for d in date_range]
        date_values = [d.date() for d in date_range]

        matrix_data = []
        for product_code in product_codes:
            product_data = progress_df[progress_df['product_code'] == product_code]
            if product_data.empty:
                continue

            product_name = name_map.get(product_code)
            if not product_name:
                product_name = product_data['product_name'].iloc[0] if 'product_name' in product_data.columns else ''
            row = {
                'è£½å“ã‚³ãƒ¼ãƒ‰': product_code,
                'è£½å“å': product_name
            }

            for date_obj, date_str in zip(date_values, date_columns):
                day_data = product_data[product_data['delivery_date'] == date_obj]
                if not day_data.empty:
                    planned_qty = pd.to_numeric(day_data['planned_quantity'], errors='coerce').fillna(0).sum()
                    row[date_str] = int(planned_qty) if planned_qty > 0 else 0
                else:
                    row[date_str] = 0

            matrix_data.append(row)

        matrix_df = pd.DataFrame(matrix_data)
        if not matrix_df.empty:
            matrix_df = matrix_df.set_index('è£½å“ã‚³ãƒ¼ãƒ‰')

        return matrix_df

    def _export_internal_orders_to_excel(self, matrix_df: pd.DataFrame, start_date: date, end_date: date):
        """ãƒãƒˆãƒªã‚¯ã‚¹ãƒ‡ãƒ¼ã‚¿ã‚’Excelã«å‡ºåŠ›"""
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            matrix_df.to_excel(writer, sheet_name='ç¤¾å†…æ³¨æ–‡', index=True)

            workbook = writer.book
            worksheet = writer.sheets['ç¤¾å†…æ³¨æ–‡']

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                           min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = border
                    if isinstance(cell.value, (int, float)) and cell.column > 2:
                        cell.number_format = '#,##0'

            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 30
            for col_idx in range(3, worksheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 12

            worksheet.insert_rows(1)
            worksheet['A1'] = f"ç¤¾å†…æ³¨æ–‡ ãƒãƒˆãƒªã‚¯ã‚¹ï¼ˆç©è¼‰æ—¥: {start_date.strftime('%Y/%m/%d')} ï½ {end_date.strftime('%Y/%m/%d')}ï¼‰"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=worksheet.max_column)
            worksheet['A1'].alignment = center_alignment

        output.seek(0)
        return output
