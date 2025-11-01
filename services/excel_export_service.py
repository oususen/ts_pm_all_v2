# app/services/excel_export_service.py
from datetime import datetime
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxlがインストールされていません。pip install openpyxl を実行してください")
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from typing import Dict, Any

class ExcelExportService:
    """Excel出力サービス"""
    
    def export_loading_plan(self, plan_result: Dict[str, Any]) -> BytesIO:
        """
        積載計画をExcelファイルとして出力
        
        Args:
            plan_result: 積載計画データ
        
        Returns:
            BytesIO: Excelファイルのバイナリストリーム
        """
        wb = Workbook()
        
        # サマリーシート
        self._create_summary_sheet(wb, plan_result)
        
        # 日別計画シート
        self._create_daily_plan_sheets(wb, plan_result)
        
        # 警告シート
        self._create_warnings_sheet(wb, plan_result)
        
        # 未積載アイテムシート
        self._create_unloaded_sheet(wb, plan_result)
        
        # デフォルトシートを削除
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # BytesIOに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_summary_sheet(self, wb: Workbook, plan_result: Dict):
        """サマリーシート作成"""
        ws = wb.active
        ws.title = "計画サマリー"
        
        summary = plan_result['summary']
        period = plan_result['period']
        
        # ヘッダースタイル
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # タイトル
        ws['A1'] = "積載計画サマリー"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # 計画情報
        ws['A3'] = "計画期間"
        ws['B3'] = period
        ws['A4'] = "作成日時"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # サマリー情報
        data = [
            ['項目', '値', '単位', '備考'],
            ['計画日数', summary['total_days'], '日', ''],
            ['総便数', summary['total_trips'], '便', ''],
            ['警告数', summary['total_warnings'], '件', ''],
            ['未積載数', summary['unloaded_count'], '件', ''],
            ['ステータス', summary['status'], '', '']
        ]
        
        for row_idx, row_data in enumerate(data, start=6):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 6:  # ヘッダー行
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
    
    def _create_daily_plan_sheets(self, wb: Workbook, plan_result: Dict):
        """日別計画シート作成"""
        daily_plans = plan_result['daily_plans']
        
        for date_str in sorted(daily_plans.keys()):
            plan = daily_plans[date_str]
            
            # シート名（日付形式）
            sheet_name = f"計画_{date_str.replace('-', '')}"
            ws = wb.create_sheet(title=sheet_name)
            
            # ヘッダー
            ws['A1'] = f"{date_str} の積載計画"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:H1')
            
            ws['A2'] = f"総便数: {plan['total_trips']}便"
            
            # テーブルヘッダー
            headers = ['便', 'トラック名', '製品コード', '製品名', '容器数', '合計数量', '納期', '体積率', '重量率']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # データ行
            row_idx = 5
            for truck_idx, truck_plan in enumerate(plan['trucks'], start=1):
                truck_name = truck_plan['truck_name']
                utilization = truck_plan['utilization']
                
                for item in truck_plan['loaded_items']:
                    ws.cell(row=row_idx, column=1, value=truck_idx)
                    ws.cell(row=row_idx, column=2, value=truck_name)
                    ws.cell(row=row_idx, column=3, value=item.get('product_code', ''))
                    ws.cell(row=row_idx, column=4, value=item.get('product_name', ''))
                    ws.cell(row=row_idx, column=5, value=item.get('num_containers', 0))
                    ws.cell(row=row_idx, column=6, value=item.get('total_quantity', 0))
                    
                    delivery_date = item.get('delivery_date')
                    if delivery_date:
                        if hasattr(delivery_date, 'strftime'):
                            ws.cell(row=row_idx, column=7, value=delivery_date.strftime('%Y-%m-%d'))
                        else:
                            ws.cell(row=row_idx, column=7, value=str(delivery_date))
                    
                    ws.cell(row=row_idx, column=8, value=f"{utilization['volume_rate']}%")
                    
                    row_idx += 1
            
            # 列幅調整
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col].width = 15
    
    def _create_warnings_sheet(self, wb: Workbook, plan_result: Dict):
        """警告シート作成"""
        ws = wb.create_sheet(title="警告一覧")
        
        ws['A1'] = "警告・注意事項"
        ws['A1'].font = Font(bold=True, size=14)
        
        headers = ['日付', '警告内容']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)
        
        row_idx = 4
        daily_plans = plan_result['daily_plans']
        
        for date_str in sorted(daily_plans.keys()):
            plan = daily_plans[date_str]
            
            for warning in plan.get('warnings', []):
                ws.cell(row=row_idx, column=1, value=date_str)
                ws.cell(row=row_idx, column=2, value=warning)
                row_idx += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80
    
    def _create_unloaded_sheet(self, wb: Workbook, plan_result: Dict):
        """未積載アイテムシート作成"""
        ws = wb.create_sheet(title="未積載アイテム")
        
        ws['A1'] = "積載できなかったアイテム"
        ws['A1'].font = Font(bold=True, size=14)
        
        headers = ['製品コード', '製品名', '容器数', '合計数量', '納期', '理由']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.font = Font(bold=True)
        
        row_idx = 4
        unloaded_tasks = plan_result.get('unloaded_tasks', [])
        
        for task in unloaded_tasks:
            ws.cell(row=row_idx, column=1, value=task.get('product_code', ''))
            ws.cell(row=row_idx, column=2, value=task.get('product_name', ''))
            ws.cell(row=row_idx, column=3, value=task.get('num_containers', 0))
            ws.cell(row=row_idx, column=4, value=task.get('total_quantity', 0))
            
            delivery_date = task.get('delivery_date')
            if delivery_date:
                if hasattr(delivery_date, 'strftime'):
                    ws.cell(row=row_idx, column=5, value=delivery_date.strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=row_idx, column=5, value=str(delivery_date))
            
            ws.cell(row=row_idx, column=6, value="積載容量不足")
            row_idx += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20